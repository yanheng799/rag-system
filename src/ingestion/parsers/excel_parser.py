"""Excel (.xlsx) 文档解析器（基于 openpyxl）"""

from __future__ import annotations

import logging

from openpyxl import load_workbook

from src.ingestion.parsers.base import BaseParser, ParsedElement, ParseError, format_rows

logger = logging.getLogger(__name__)

# 每个 Element 的最大行数，避免单条 element 过长
MAX_ROWS_PER_ELEMENT = 100


class ExcelParser(BaseParser):
    """使用 openpyxl 解析 Excel 文档，按 sheet 逐块提取数据"""

    def parse(self, file_path: str) -> list[ParsedElement]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise ParseError(file_path, str(e)) from e

        elements: list[ParsedElement] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            sheet_elements = self._parse_sheet(ws, sheet_idx, sheet_name)
            elements.extend(sheet_elements)

        wb.close()
        logger.info("Excel 解析完成: %s, 共 %d 个元素", file_path, len(elements))
        return elements

    def _parse_sheet(self, ws, sheet_idx: int, sheet_name: str) -> list[ParsedElement]:
        """解析单个 Sheet"""
        elements: list[ParsedElement] = []

        # 读取所有非空行
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            # 跳过全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            rows_data.append([str(cell) if cell is not None else "" for cell in row])

        if not rows_data:
            return elements

        # 第一行作为表头
        headers = rows_data[0]

        # 检测合并单元格区域（用于样式标记）
        merged_ranges = []
        if hasattr(ws, "merged_cells") and ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                merged_ranges.append(str(merged_range))

        # 按行数分组为多个 ParsedElement
        data_rows = rows_data[1:]

        for chunk_index, i in enumerate(range(0, len(data_rows), MAX_ROWS_PER_ELEMENT)):
            chunk_rows = data_rows[i : i + MAX_ROWS_PER_ELEMENT]
            content = format_rows(headers, chunk_rows, sheet_name)

            elements.append(
                ParsedElement(
                    elem_type="table",
                    content=content,
                    page=sheet_idx,
                    bbox=(0, chunk_index, 0, chunk_index + 1),
                    style={
                        "sheet_name": sheet_name,
                        "headers": headers,
                        "has_merged_cells": len(merged_ranges) > 0,
                    },
                    raw={"headers": headers, "rows": chunk_rows},
                )
            )

        return elements

    def supported_types(self) -> list[str]:
        return ["xlsx"]
